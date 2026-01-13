"""
Reusable Excel formatting utilities for pivot tables.
"""

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

class ExcelFormatter:
    """
    Handles all aesthetic formatting for Excel worksheets.
    """
    
    def __init__(self):
        # Define standard colors
        self.blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        self.white_bold_font = Font(bold=True, color="FFFFFF")
        self.bold_font = Font(bold=True)
    
    def apply_standard_formatting(self, worksheet, dataframe, grand_total_identifier='Grand Total', 
                                   bold_columns=None, header_row=1, data_start_row=2):
        """
        Args:
            worksheet: openpyxl worksheet object
            dataframe: pandas DataFrame that was written to the worksheet
            grand_total_identifier: String to identify Grand Total rows (default: 'Grand Total')
            bold_columns: List of column names to make bold (default: None)
            header_row: Row number of headers (default: 1)
            data_start_row: First row of data (default: 2)
        """
        
        # Format header row
        self._format_header_row(worksheet, dataframe, header_row)
        
        # Enable autofilter on the header row
        self._enable_filters(worksheet)
        
        # Format Grand Total rows and bold columns
        self._format_data_rows(worksheet, dataframe, grand_total_identifier, 
                              bold_columns, data_start_row)
        
        self.auto_adjust_column_widths(worksheet, dataframe)

    
    def _format_header_row(self, worksheet, dataframe, header_row):
        """
        Format the header row with blue background and white bold text.
        """
        
        for col_num in range(1, len(dataframe.columns) + 1):
            cell = worksheet.cell(row=header_row, column=col_num)
            cell.fill = self.blue_fill
            cell.font = self.white_bold_font
    
    def _enable_filters(self, worksheet):
        """
        Enable Excel autofilter on all columns.
        """

        worksheet.auto_filter.ref = worksheet.dimensions
    
    def _format_data_rows(self, worksheet, dataframe, grand_total_identifier, 
                         bold_columns, data_start_row):
        """
        Format data rows including Grand Total rows and bold columns.
        """

        # Find indices of columns to bold
        bold_col_indices = []
        if bold_columns:
            for col_name in bold_columns:
                if col_name in dataframe.columns:
                    # +1 because Excel columns are 1-indexed
                    bold_col_indices.append(list(dataframe.columns).index(col_name) + 1)
        
        # Loop through all data rows
        num_rows = len(dataframe)
        for row_num in range(data_start_row, num_rows + data_start_row):
            # Check if this is a Grand Total row
            # Check the third column (Media Name or Unit typically)
            third_col_value = worksheet.cell(row=row_num, column=3).value
            
            is_grand_total = (third_col_value == grand_total_identifier)
            
            if is_grand_total:
                # Format entire Grand Total row with blue background and white bold text
                for col_num in range(1, len(dataframe.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = self.orange_fill
                    cell.font = self.white_bold_font
            else:
                # Bold specific columns for regular data rows
                for col_idx in bold_col_indices:
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    cell.font = self.bold_font
    
    def auto_adjust_column_widths(self, worksheet, dataframe, min_width=8, max_width=40):
        """
        Auto-adjust column widths based on content.
        """

        for col_num, column in enumerate(dataframe.columns, start=1):
            column_letter = get_column_letter(col_num)
            
            # Calculate max length in column
            max_length = len(str(column))  # Start with header length
            
            for row_num in range(2, len(dataframe) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width with constraints
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width