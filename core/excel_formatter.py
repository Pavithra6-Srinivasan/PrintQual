"""
Excel formatter
"""

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

class ExcelFormatter:
    """Handles all aesthetic formatting."""
    
    def __init__(self):
        # Define standard colors
        self.blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        self.white_bold_font = Font(bold=True, color="FFFFFF")
        self.bold_font = Font(bold=True)
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    
    def apply_standard_formatting(self, worksheet, dataframe, grand_total_identifier='Grand Total', 
                                   bold_columns=None, header_row=1, data_start_row=2, 
                                   highlight_threshold=None, total_column_name=None):
        
        self.format_header_row(worksheet, dataframe, header_row)
        self.enable_filters(worksheet)
        self.format_data_rows(worksheet, dataframe, grand_total_identifier, bold_columns, data_start_row)
        self.format_Result_column(worksheet, dataframe, data_start_row)

        if highlight_threshold is not None:
            self.highlight_high_error_cells(worksheet, dataframe, highlight_threshold,
                                            grand_total_identifier, data_start_row, 
                                            total_column_name=total_column_name)
        
        self.auto_adjust_column_widths(worksheet, dataframe)
    
    def format_header_row(self, worksheet, dataframe, header_row):
        """Format the header row with blue background and white bold text."""
        for col_num in range(1, len(dataframe.columns) + 1):
            cell = worksheet.cell(row=header_row, column=col_num)
            cell.fill = self.blue_fill
            cell.font = self.white_bold_font
    
    def enable_filters(self, worksheet):
        """Enable Excel autofilter on all columns."""
        worksheet.auto_filter.ref = worksheet.dimensions
    
    def format_data_rows(self, worksheet, dataframe, grand_total_identifier, 
                         bold_columns, data_start_row):
        """
        Bold column headings and format data rows including Grand Total rows.
        """
        # Find indices of columns to bold
        bold_col_indices = []
        if bold_columns:
            for col_name in bold_columns:
                if col_name in dataframe.columns:
                    bold_col_indices.append(list(dataframe.columns).index(col_name) + 1)
        
        grand_total_col_idx = None

        for col_name in dataframe.columns:
            if dataframe[col_name].astype(str).str.contains(grand_total_identifier).any():
                grand_total_col_idx = list(dataframe.columns).index(col_name) + 1
                break
        
        # Loop through all data rows
        num_rows = len(dataframe)
        for row_num in range(data_start_row, num_rows + data_start_row):
            # Check if this is a Grand Total row
            is_grand_total = False
            if grand_total_col_idx:
                cell_value = worksheet.cell(row=row_num, column=grand_total_col_idx).value
                is_grand_total = (str(cell_value) == grand_total_identifier)
            
            if is_grand_total:
                # Format entire Grand Total row with orange background and white bold text
                for col_num in range(1, len(dataframe.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = self.orange_fill
                    cell.font = self.white_bold_font
            else:
                # Bold specific columns for regular data rows
                for col_idx in bold_col_indices:
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    cell.font = self.bold_font
    
    def auto_adjust_column_widths(self, worksheet, dataframe, min_width=8, max_width=20):
        """Auto-adjust column widths based on content."""
        for col_num, column in enumerate(dataframe.columns, start=1):
            column_letter = get_column_letter(col_num)
            
            # Calculate max length in column
            max_length = len(str(column))  # Start with header length
            
            for row_num in range(2, len(dataframe) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width with constraints
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def format_Result_column(self, worksheet, dataframe, data_start_row=2):
        """
        Color Result column cells: green for PASS, red for FAIL.
        """
        # Find Result column index
        result_col_idx = None
        for col_num, col_name in enumerate(dataframe.columns, start=1):
            if col_name == 'Result':
                result_col_idx = col_num
                break

        if result_col_idx is None:
            return

        # Color each cell based on PASS/FAIL value
        for row_num in range(data_start_row, len(dataframe) + data_start_row):
            cell = worksheet.cell(row=row_num, column=result_col_idx)
            cell_value = str(cell.value).upper() if cell.value else ""
            
            if cell_value == 'PASS':
                cell.fill = self.green_fill
            elif cell_value == 'FAIL':
                cell.fill = self.red_fill

    def highlight_high_error_cells(self, worksheet, dataframe, threshold, 
                                    grand_total_identifier, data_start_row,
                                    total_column_name=None, use_relative_threshold=True):
        """
        Highlight individual error cells that contribute significantly to failures.
        """
        highlight_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        
        grand_total_col_idx = None

        for col_name in dataframe.columns:
            if dataframe[col_name].astype(str).str.contains(grand_total_identifier).any():
                grand_total_col_idx = list(dataframe.columns).index(col_name) + 1
                break

        # Identify total per-K column
        total_per_k_col_idx = None
        
        if total_column_name and total_column_name in dataframe.columns:
            total_per_k_col_idx = list(dataframe.columns).index(total_column_name) + 1
        else:
            return
        
        # Find all per-K columns (excluding the total sum column)
        per_k_columns = []
        for col_num, col_name in enumerate(dataframe.columns, start=1):
            col_str = str(col_name)
            if col_str.endswith('/K') and col_num != total_per_k_col_idx:   
                per_k_columns.append((col_num, col_name))
        
        # Highlight cells
        for row_num in range(data_start_row, len(dataframe) + data_start_row):
            # Skip Grand Total rows
            if grand_total_col_idx:
                label = worksheet.cell(row=row_num, column=grand_total_col_idx).value
                if str(label) == grand_total_identifier:
                    continue

            # Get threshold for this row
            if use_relative_threshold:
                try:
                    total_cell = worksheet.cell(row=row_num, column=total_per_k_col_idx)
                    row_total = float(total_cell.value or 0.0)
                    if row_total <= 0:
                        continue
                    row_threshold = row_total * 0.5
                except (ValueError, TypeError):
                    continue
            else:
                row_threshold = threshold

            # Check each error column
            for col_num, col_name in per_k_columns:
                cell = worksheet.cell(row=row_num, column=col_num)
                try:
                    value = float(cell.value or 0.0)
                    if value > row_threshold:
                        cell.fill = highlight_fill
                except (ValueError, TypeError):
                    pass