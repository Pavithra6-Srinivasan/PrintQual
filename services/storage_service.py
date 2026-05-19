import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Shared style constants ────────────────────────────────────────────────────
_LIGHT_BLUE = "BDD7EE"   # header row
_BLUE       = "4472C4"   # Total rows
_GREEN      = "C6EFCE"
_RED        = "FF9999"
_WHITE      = "FFFFFF"
_BLACK      = "000000"
_DARK_BLUE  = "1F3864"   # header text

_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Fixed display order for dimension columns
_DIM_FIELD_ORDER = [
    "Test Condition",
    "Media Type",
    "Media Cat",
    "Input Tray", "Input_Tray", "Tray",
    "Print Mode",
]

CATEGORY_ORDER = [
    "Intervention", "Soft Error", "Skew", "Other Defects",
    "PQ", "Image Quality", "Other Issue",
]


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color=_BLACK, bold=False, size=9):
    return Font(color=hex_color, bold=bold, size=size)


class StorageService:

    def save_full_report(self, output_path, summary_data, all_pivots):
        # Three sheet dicts — all written via openpyxl, interleaved per category
        media_sheets  = {}   # category_name[:31]            -> (df, flags)
        unit_sheets   = {}   # "cat by Unit"[:31]            -> (df, flags)
        detail_sheets = {}   # "cat Detail"[:31]             -> (df, flags)

        for category_name, pivot_data in all_pivots.items():
            config        = pivot_data["config"]
            combined_df   = pivot_data["combined"].copy()
            spec_has_tray = pivot_data.get("spec_has_tray", True)
            sub_assembly  = str(pivot_data.get("sub_assembly", "")).strip().upper()

            # Drop Input Tray when spec has no tray entries or sub-assembly is ADF
            if not spec_has_tray or sub_assembly == "ADF":
                tray_col = next(
                    (c for c in ("Input Tray", "Input_Tray", "Tray")
                     if c in combined_df.columns), None
                )
                if tray_col:
                    combined_df = combined_df.drop(columns=[tray_col])

            # Drop Grand Total rows — unit rows carry their own Spec Limit / Result
            if "Unit" in combined_df.columns:
                gt_mask = (
                    combined_df["Unit"].astype(str).str.strip().str.lower()
                    == "grand total"
                )
                unit_df = combined_df[~gt_mask].copy()
            else:
                unit_df = combined_df.copy()

            if "Spec Limit" in unit_df.columns:
                unit_df["Spec Limit"] = pd.to_numeric(
                    unit_df["Spec Limit"], errors="coerce"
                )

            # Category has a spec only if at least one row is PASS or FAIL
            has_spec = (
                "Result" in unit_df.columns
                and unit_df["Result"].isin({"PASS", "FAIL"}).any()
            )

            # Reverse-compute raw counts from /K rates for weighted averages
            denom_col = getattr(config, 'denominator_column', 'Tpages')
            if denom_col not in unit_df.columns:
                denom_col = 'Tpages'
            per_k_cols = [
                c for c in unit_df.columns
                if str(c).endswith("/K") and c != config.total_column_name
            ]
            per_k_info = []
            tpages_s = pd.to_numeric(
                unit_df.get(denom_col, pd.Series(0, index=unit_df.index)),
                errors="coerce"
            ).fillna(0)
            for i, k_col in enumerate(per_k_cols):
                raw_name = f"_r{i}"
                rates = pd.to_numeric(unit_df[k_col], errors="coerce").fillna(0)
                unit_df[raw_name] = rates * tpages_s / 1000
                per_k_info.append((k_col, raw_name))

            total_col = config.total_column_name

            # Pre-filter blank/NaN units for unit and detail sheets
            if "Unit" in unit_df.columns:
                u_str = unit_df["Unit"].astype(str).str.strip()
                unit_df_filt = unit_df[
                    unit_df["Unit"].notna() &
                    (u_str != "") &
                    (u_str.str.lower() != "nan")
                ].copy()
            else:
                unit_df_filt = unit_df.copy()

            # ── Sheet 1: Media Name (aggregated across units) ─────────────
            if "Media Name" in unit_df.columns:
                media_df = self._aggregate_by_media(unit_df, per_k_info, total_col, denom_col)
                df, flags = self._build_table(
                    media_df, "Media Name", per_k_info, total_col, has_spec, denom_col=denom_col
                )
                media_sheets[category_name[:31]] = (df, flags)

            # ── Sheet 2: By Unit (aggregated across media names) ──────────
            if "Unit" in unit_df.columns and not unit_df_filt.empty:
                unit_agg = self._aggregate_by_unit(unit_df_filt, per_k_info, total_col, denom_col)
                df, flags = self._build_table(
                    unit_agg, "Unit", per_k_info, total_col, has_spec, denom_col=denom_col
                )
                unit_sheets[f"{category_name} by Unit"[:31]] = (df, flags)

            # ── Sheet 3: Detail (Media Name + Unit, all combinations) ─────
            if ("Unit" in unit_df.columns and "Media Name" in unit_df.columns
                    and not unit_df_filt.empty):
                df, flags = self._build_table(
                    unit_df_filt, "Unit", per_k_info, total_col, has_spec,
                    additional_dim_cols=["Media Name"], denom_col=denom_col
                )
                detail_sheets[f"{category_name} Detail"[:31]] = (df, flags)

        # ── Interleave sheets in CATEGORY_ORDER: media → unit → detail ────
        all_sheets  = {}
        seen_unit   = set()
        seen_detail = set()

        for cat in CATEGORY_ORDER:
            if cat in media_sheets:
                all_sheets[cat] = media_sheets[cat]
            unit_sname = f"{cat} by Unit"[:31]
            if unit_sname in unit_sheets:
                all_sheets[unit_sname] = unit_sheets[unit_sname]
                seen_unit.add(unit_sname)
            detail_sname = f"{cat} Detail"[:31]
            if detail_sname in detail_sheets:
                all_sheets[detail_sname] = detail_sheets[detail_sname]
                seen_detail.add(detail_sname)

        # Categories not in CATEGORY_ORDER go at the end
        for name, data in media_sheets.items():
            if name not in all_sheets:
                all_sheets[name] = data
                unit_sname = f"{name} by Unit"[:31]
                if unit_sname in unit_sheets and unit_sname not in seen_unit:
                    all_sheets[unit_sname] = unit_sheets[unit_sname]
                detail_sname = f"{name} Detail"[:31]
                if detail_sname in detail_sheets and detail_sname not in seen_detail:
                    all_sheets[detail_sname] = detail_sheets[detail_sname]

        # ── Write all sheets with openpyxl ────────────────────────────────
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for name, (df, _flags) in all_sheets.items():
                df.to_excel(writer, sheet_name=name, index=False)

            wb = writer.book
            for name, (df, flags) in all_sheets.items():
                if name in wb.sheetnames:
                    self._format_sheet(wb[name], list(df.columns), flags)

    # -------------------------------------------------------------------------

    def _aggregate_by_media(self, unit_df, per_k_info, total_col=None, denom_col='Tpages'):
        """
        Collapse Unit so each Media Name appears once per dimension group.
        Denominator and raw counts summed; Spec Limit = first valid; Result = FAIL if any.
        """
        if "Unit" not in unit_df.columns:
            return unit_df

        raw_names = {rn for _, rn in per_k_info}
        k_names   = {kn for kn, _ in per_k_info}
        exclude   = {"Unit", denom_col} | raw_names | k_names | {"Spec Limit", "Result"}
        if total_col:
            exclude.add(total_col)
        group_cols = [c for c in unit_df.columns if c not in exclude]

        agg_dict = {}
        if denom_col in unit_df.columns:
            agg_dict[denom_col] = "sum"
        for rn in raw_names:
            if rn in unit_df.columns:
                agg_dict[rn] = "sum"
        if "Spec Limit" in unit_df.columns:
            agg_dict["Spec Limit"] = lambda x: (
                float(x.dropna().iloc[0]) if x.notna().any() else None
            )
        if "Result" in unit_df.columns:
            agg_dict["Result"] = lambda x: (
                "FAIL" if (x == "FAIL").any() else
                ("PASS" if (x == "PASS").any() else "")
            )

        return (
            unit_df
            .groupby(group_cols, sort=False, dropna=False)
            .agg(agg_dict)
            .reset_index()
        )

    # -------------------------------------------------------------------------

    def _aggregate_by_unit(self, unit_df, per_k_info, total_col=None, denom_col='Tpages'):
        """
        Collapse Media Name so each Unit appears once per dimension group.
        Denominator and raw counts summed; Spec Limit = first valid; Result = FAIL if any.
        """
        raw_names = {rn for _, rn in per_k_info}
        k_names   = {kn for kn, _ in per_k_info}
        exclude   = {"Media Name", denom_col} | raw_names | k_names | {"Spec Limit", "Result"}
        if total_col:
            exclude.add(total_col)
        group_cols = [c for c in unit_df.columns if c not in exclude]

        agg_dict = {}
        if denom_col in unit_df.columns:
            agg_dict[denom_col] = "sum"
        for rn in raw_names:
            if rn in unit_df.columns:
                agg_dict[rn] = "sum"
        if "Spec Limit" in unit_df.columns:
            agg_dict["Spec Limit"] = lambda x: (
                float(x.dropna().iloc[0]) if x.notna().any() else None
            )
        if "Result" in unit_df.columns:
            agg_dict["Result"] = lambda x: (
                "FAIL" if (x == "FAIL").any() else
                ("PASS" if (x == "PASS").any() else "")
            )

        return (
            unit_df
            .groupby(group_cols, sort=False, dropna=False)
            .agg(agg_dict)
            .reset_index()
        )

    # -------------------------------------------------------------------------

    def _build_table(self, unit_df, leaf_col, per_k_info, total_col, has_spec,
                     additional_dim_cols=None, denom_col='Tpages'):
        """
        Build the display DataFrame for one sheet.

        For each dimension group one Total row is written first (weighted-avg
        /K rates, summed denominator) followed by individual leaf rows with dim
        values repeated on every row.  additional_dim_cols adds extra columns
        (e.g. "Media Name") to the grouping for the Detail sheet.
        """
        present = set(unit_df.columns)
        dim_cols = [c for c in _DIM_FIELD_ORDER if c in present]
        if additional_dim_cols:
            for col in additional_dim_cols:
                if col in present and col not in dim_cols:
                    dim_cols.append(col)

        k_col_names = [k for k, _ in per_k_info]

        out_cols = dim_cols + [leaf_col, denom_col] + k_col_names
        if per_k_info:
            out_cols.append(total_col)
        if "Spec Limit" in unit_df.columns:
            out_cols.append("Spec Limit")
        if has_spec:
            out_cols.append("Result")

        rows  = []
        flags = []   # True = Total (blue) row

        groups = (
            unit_df.groupby(dim_cols, sort=False, dropna=False)
            if dim_cols else [("__all__", unit_df)]
        )

        for gk, grp in groups:
            if dim_cols:
                gk_t     = gk if isinstance(gk, tuple) else (gk,)
                dim_vals = dict(zip(dim_cols, gk_t))
            else:
                dim_vals = {}

            tp_sum = pd.to_numeric(grp[denom_col], errors="coerce").fillna(0).sum() \
                     if denom_col in grp.columns else 0.0

            # ── Detail rows (dim values repeated on every row) ───────────
            for _, r in grp.iterrows():
                dr = {c: dim_vals.get(c, "") for c in dim_cols}
                dr[leaf_col] = r.get(leaf_col, "")

                tp = pd.to_numeric(r.get(denom_col, 0), errors="coerce") or 0
                dr[denom_col] = int(tp) if tp else ""

                for k_col, rn in per_k_info:
                    rv = pd.to_numeric(r.get(rn, 0), errors="coerce") or 0
                    dr[k_col] = round(rv / tp * 1000, 2) if tp else 0.0
                if per_k_info:
                    ar = sum(
                        pd.to_numeric(r.get(rn, 0), errors="coerce") or 0
                        for _, rn in per_k_info
                    )
                    dr[total_col] = round(ar / tp * 1000, 2) if tp else 0.0

                if "Spec Limit" in unit_df.columns:
                    sv = pd.to_numeric(r.get("Spec Limit"), errors="coerce")
                    dr["Spec Limit"] = float(sv) if pd.notna(sv) else None

                if has_spec:
                    sl_val   = dr.get("Spec Limit")
                    rate_val = dr.get(total_col, 0.0) or 0.0
                    if sl_val is not None and pd.notna(sl_val):
                        dr["Result"] = "PASS" if rate_val <= float(sl_val) else "FAIL"
                    elif "Result" in unit_df.columns:
                        dr["Result"] = r.get("Result", "")
                    else:
                        dr["Result"] = ""

                rows.append({c: dr.get(c, "") for c in out_cols})
                flags.append(False)

            # ── Total row at bottom of group ─────────────────────────────
            tr = {**dim_vals, leaf_col: "Total"}
            tr[denom_col] = int(round(tp_sum)) if tp_sum else 0

            for k_col, rn in per_k_info:
                rs = pd.to_numeric(grp[rn], errors="coerce").fillna(0).sum()
                tr[k_col] = round(rs / tp_sum * 1000, 2) if tp_sum else 0.0
            if per_k_info:
                all_rs = sum(
                    pd.to_numeric(grp[rn], errors="coerce").fillna(0).sum()
                    for _, rn in per_k_info
                )
                tr[total_col] = round(all_rs / tp_sum * 1000, 2) if tp_sum else 0.0

            if "Spec Limit" in unit_df.columns:
                sv = pd.to_numeric(grp["Spec Limit"], errors="coerce").dropna()
                tr["Spec Limit"] = float(sv.iloc[0]) if len(sv) else None

            if has_spec:
                sl_val   = tr.get("Spec Limit")
                rate_val = tr.get(total_col, 0.0) or 0.0
                if sl_val is not None and pd.notna(sl_val):
                    tr["Result"] = "PASS" if rate_val <= float(sl_val) else "FAIL"
                elif "Result" in unit_df.columns:
                    tr["Result"] = "FAIL" if (grp["Result"] == "FAIL").any() else "PASS"
                else:
                    tr["Result"] = ""

            rows.append({c: tr.get(c, "") for c in out_cols})
            flags.append(True)

        return pd.DataFrame(rows, columns=out_cols), flags

    # -------------------------------------------------------------------------

    def _format_sheet(self, ws, col_names, is_total_flags):
        n_cols = len(col_names)

        spec_ci   = next((i + 1 for i, h in enumerate(col_names)
                          if str(h).lower() == "spec limit"), None)
        result_ci = next((i + 1 for i, h in enumerate(col_names)
                          if str(h).lower() == "result"), None)
        k_cis     = [i + 1 for i, h in enumerate(col_names)
                     if str(h).endswith("/K") or str(h).lower().startswith("sum of")]

        # ── Header row ───────────────────────────────────────────────────
        for ci in range(1, n_cols + 1):
            c = ws.cell(1, ci)
            c.fill      = _fill(_LIGHT_BLUE)
            c.font      = _font(_DARK_BLUE, bold=True)
            c.alignment = Alignment(
                text_rotation=90, vertical="bottom", wrap_text=False
            )
            c.border = _BORDER
        ws.row_dimensions[1].height = 80

        # ── Data rows ────────────────────────────────────────────────────
        for ri, is_total in enumerate(is_total_flags, start=2):
            for ci in range(1, n_cols + 1):
                cell = ws.cell(ri, ci)
                cell.border    = _BORDER
                cell.alignment = Alignment(vertical="center")
                if is_total:
                    cell.fill = _fill(_BLUE)
                    cell.font = _font(_WHITE, bold=True)

            if result_ci:
                rc = ws.cell(ri, result_ci)
                if rc.value == "PASS":
                    rc.fill = _fill(_GREEN)
                    rc.font = _font(_BLACK, bold=True)
                    rc.value = "No Issue"
                elif rc.value == "FAIL":
                    rc.fill = _fill(_RED)
                    rc.font = _font(_BLACK, bold=True)
                    rc.value = "With observation"

            # Highlight individual /K values that exceed spec limit
            if spec_ci and k_cis:
                try:
                    spec_val = float(ws.cell(ri, spec_ci).value)
                except (TypeError, ValueError):
                    spec_val = None
                if spec_val is not None:
                    for kci in k_cis:
                        try:
                            kv = float(ws.cell(ri, kci).value)
                        except (TypeError, ValueError):
                            continue
                        if kv > spec_val:
                            ws.cell(ri, kci).fill = _fill(_RED)

        if spec_ci:
            for ri in range(2, len(is_total_flags) + 2):
                ws.cell(ri, spec_ci).number_format = "0.00"

        # ── AutoFilter on full table ──────────────────────────────────────
        n_rows = len(is_total_flags) + 1
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"

        # ── Column widths ─────────────────────────────────────────────────
        for ci, h in enumerate(col_names, start=1):
            h_str = str(h)
            ltr   = get_column_letter(ci)
            if h_str.endswith("/K") or h_str.lower().startswith("sum of"):
                ws.column_dimensions[ltr].width = 8
            elif h_str.lower() in ("tpages", "spec limit", "result"):
                ws.column_dimensions[ltr].width = 8
            elif h_str.lower() in ("media name", "unit"):
                ws.column_dimensions[ltr].width = 30
            else:
                ws.column_dimensions[ltr].width = 14

    # -------------------------------------------------------------------------
    # DATABASE SAVE
    # -------------------------------------------------------------------------

    def save_results_to_database(self, db, all_pivots, year, quarter, summary_data):
        for cat in summary_data["categories"]:
            category_name = cat["category"]
            for media in cat["media_summaries"]:
                media_type     = media["media_type"]
                overall_result = media["overall_result"]
                errors         = media.get("errors", [])

                all_media_names  = set()
                all_error_types  = set()
                failed_units_set = set()

                for err in errors:
                    all_error_types.add(f"{err['error']} ({err['rate']:.3f}/K)")
                    for me in err.get("failed_media", []):
                        all_media_names.add(me["media_name"])
                        failed_units_set.update(me["units"])

                db.insert_summary_result(
                    category=category_name,
                    media_type=media_type,
                    overall_result=overall_result,
                    failed_units=",".join(sorted(failed_units_set)),
                    failed_media_names=",".join(sorted(all_media_names)),
                    failed_error_types=",".join(sorted(all_error_types)),
                    failed_conditions="",
                    remarks=self._build_remarks(errors, overall_result),
                    year=year,
                    quarter=quarter
                )

    def _build_remarks(self, errors, overall_result):
        if not errors or overall_result.upper() != "FAIL":
            return ""
        parts = []
        for err in errors:
            media_parts = [
                f"{me['media_name']} ({', '.join(me['units'])})"
                for me in err["failed_media"]
            ]
            parts.append(
                f"{err['error']} {err['rate']:.3f}/K: {' | '.join(media_parts)}"
            )
        return " // ".join(parts)
